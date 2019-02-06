from typing import Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from data import CompanyData, Field

ticker_font = Font(name='Calibri',
                   size=14,
                   bold=True,
                   italic=False,
                   vertAlign=None,
                   underline='none',
                   strike=False,
                   color='FF000000')
ticker_alignment = Alignment(horizontal='center',
                             vertical='bottom',
                             text_rotation=0,
                             wrap_text=False,
                             shrink_to_fit=False,
                             indent=0)

field_font = Font(name='Calibri',
                  size=12,
                  bold=True,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')
field_alignment = Alignment(horizontal='center',
                            vertical='bottom',
                            text_rotation=0,
                            wrap_text=False,
                            shrink_to_fit=False,
                            indent=0)


def adjust_columns(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width


def generate_reports(output_path: str, companies: Dict[str, CompanyData]):
    workbook = Workbook()
    sheet = workbook.create_sheet("Companies")
    row = 1

    for ticker, company_data in companies.items():
        sheet.merge_cells(f'A{row}:F{row}')
        sheet.cell(row, 1).value = ticker
        sheet.cell(row, 1).font = ticker_font
        sheet.cell(row, 1).alignment = ticker_alignment
        row += 1
        column = 2

        sheet.cell(row + 1, 1).value = '10 year'
        sheet.cell(row + 1, 1).font = field_font
        sheet.cell(row + 2, 1).value = '5 year'
        sheet.cell(row + 2, 1).font = field_font
        sheet.cell(row + 3, 1).value = '1 year'
        sheet.cell(row + 3, 1).font = field_font
        # sheet.cell(row + 3, 1).s

        for field in Field:
            indicator = company_data.get_indicator(field)

            sheet.cell(row, column).value = field.value.replace('_', ' ')
            sheet.cell(row, column).font = field_font
            sheet.cell(row, column).alignment = field_alignment
            sheet.cell(row + 1, column).value = round(indicator.year10 * 100, 2)
            sheet.cell(row + 2, column).value = round(indicator.year5 * 100, 2)
            sheet.cell(row + 3, column).value = round(indicator.year1 * 100, 2)

            column += 1

        row += 5

    default_sheet = workbook.get_sheet_by_name('Sheet')
    workbook.remove_sheet(default_sheet)
    adjust_columns(sheet)
    workbook.save(output_path)
